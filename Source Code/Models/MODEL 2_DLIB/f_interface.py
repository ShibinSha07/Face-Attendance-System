import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import face_recognition
import cv2
from datetime import datetime
import os
import threading
import openpyxl  # Import openpyxl for working with Excel files

class FaceRecognitionApp:
    EXCEL_FILE_PATH = r"C:\Users\rohan\OneDrive\Documents\student_file\student_name.xlsx"
    KNOWN_FACES_FOLDER = r"C:\Users\rohan\OneDrive\Pictures\Screenshots"
    ATTENDANCE_HEADER = "Attendance"
    STUDENT_NAME_HEADER = "Student Name"

    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition Attendance System")

        # Initialize face recognition variables
        self.known_face_encodings = []
        self.known_face_names = []
        self.attendance_status = {}

        # Load icons
        self.start_icon = self.load_and_resize_icon("start.png", (20, 20))
        self.mark_icon = self.load_and_resize_icon("mark.png", (20, 20))
        self.list_icon = self.load_and_resize_icon("list.png", (20, 20))
        self.exit_icon = self.load_and_resize_icon("exit.png", (20, 20))

        # Create GUI components
        self.create_gui()

    def create_gui(self):
        # Create GUI components
        self.label = tk.Label(self.root, text="Face Recognition Attendance System", font=("Helvetica", 16))
        self.label.pack(pady=10)

        # Create buttons
        self.start_button = tk.Button(self.root, text="Start Face Recognition", command=self.start_recognition_thread, compound=tk.LEFT, image=self.start_icon)
        self.start_button.pack(pady=10)

        self.mark_attendance_button = tk.Button(self.root, text="Mark Attendance", command=self.mark_attendance, compound=tk.LEFT, image=self.mark_icon)
        self.mark_attendance_button.pack(pady=10)

        self.student_list_button = tk.Button(self.root, text="Student List", command=self.show_student_list, compound=tk.LEFT, image=self.list_icon)
        self.student_list_button.pack(pady=10)

        self.exit_button = tk.Button(self.root, text="Exit", command=self.exit_app, compound=tk.LEFT, image=self.exit_icon)
        self.exit_button.pack(pady=10)

    def load_and_resize_icon(self, path, size):
        original_icon = Image.open(path)
        resized_icon = original_icon.resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(resized_icon)

    def load_known_faces(self):
        # Iterate over each file in the known_faces folder
        for filename in os.listdir(self.KNOWN_FACES_FOLDER):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                # Construct the full file path
                known_image_path = os.path.join(self.KNOWN_FACES_FOLDER, filename)

                # Load the known face image
                known_image = face_recognition.load_image_file(known_image_path)

                # Find face locations in the image
                face_locations = face_recognition.face_locations(known_image)

                # If a face is found, extract the face encoding
                if face_locations:
                    # Assume there's only one face in the image for simplicity
                    face_encoding = face_recognition.face_encodings(known_image, face_locations)[0]

                    # Add the face encoding and corresponding name to the lists
                    self.known_face_encodings.append(face_encoding)
                    self.known_face_names.append(os.path.splitext(filename)[0])  # Use the file name as the person's name

                    # Initialize attendance status to False for each known face
                    self.attendance_status[os.path.splitext(filename)[0]] = False
                else:
                    print(f"No face found in {known_image_path}")

    def start_recognition_thread(self):
        # Start a new thread for face recognition
        recognition_thread = threading.Thread(target=self.start_recognition)
        recognition_thread.start()

    def start_recognition(self):
        # Initialize video capture
        self.video_capture = cv2.VideoCapture(0)

        # Load known faces
        self.load_known_faces()

        while True:
            # Capture frame-by-frame
            ret, frame = self.video_capture.read()

            # Find all face locations and face encodings in the current frame
            face_locations = face_recognition.face_locations(frame)
            face_encodings = face_recognition.face_encodings(frame, face_locations)

            # Loop through each face found in the frame
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                # Compare the current face with known faces
                matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding)

                name = "Unknown"

                # If a match is found, use the name of the known face
                if True in matches:
                    first_match_index = matches.index(True)
                    name = self.known_face_names[first_match_index]

                    # Check if attendance has already been marked for this face
                    if not self.attendance_status[name]:
                        # Mark attendance and update the status
                        self.attendance_status[name] = True
                        attendance_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        print(f"Attendance Marked for {name} at {attendance_time}")

                        # Call a function to mark attendance in the Excel file
                        self.mark_attendance_in_excel(name, attendance_time)

                # Draw a rectangle around the face and display the name
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

            # Display the resulting frame
            cv2.imshow('Video', frame)

            # Break the loop when 'q' key is pressed
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        # Release the webcam and close all windows
        self.video_capture.release()
        cv2.destroyAllWindows()

    def mark_attendance_in_excel(self, student_name, attendance_time):
        try:
            # Load the existing workbook or create a new one if it doesn't exist
            wb = self.load_or_create_workbook()

            sheet = wb.active

            # Iterate through rows to find the student
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] == student_name:
                    # Find the current date column index or add a new column
                    current_date = datetime.now().strftime("%Y-%m-%d")
                    if current_date not in sheet[1]:
                        sheet.append([current_date])  # Add a new column for the current date

                    date_column = sheet[1].index(current_date) + 1  # Column index is 1-based
                    # Mark attendance for the student
                    sheet.cell(row=sheet.iter_rows(min_row=2, max_col=1).index(row) + 2, column=date_column, value=attendance_time)
                    break

            # Save the workbook
            wb.save(self.EXCEL_FILE_PATH)
        except Exception as e:
            print(f"Error marking attendance in Excel: {e}")

    def load_or_create_workbook(self):
        try:
            # Load the existing workbook or create a new one
            if os.path.exists(self.EXCEL_FILE_PATH):
                wb = openpyxl.load_workbook(self.EXCEL_FILE_PATH)
            else:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append([self.STUDENT_NAME_HEADER])
                wb.save(self.EXCEL_FILE_PATH)

            return wb
        except Exception as e:
            print(f"Error loading or creating workbook: {e}")

    def mark_attendance(self):
        # Placeholder for attendance marking logic
        # You can implement specific logic for manual attendance marking if needed
        pass

    def show_student_list(self):
        # Create a new window for the student list
        student_list_window = tk.Toplevel(self.root)
        student_list_window.title("Student List")

        # Create a listbox to display the student names
        listbox = tk.Listbox(student_list_window, selectmode=tk.MULTIPLE)
        for student_name in self.known_face_names:
            listbox.insert(tk.END, student_name)
        listbox.pack(padx=20, pady=20)

        # Create a button to mark attendance for selected students
        mark_button = tk.Button(student_list_window, text="Mark Attendance", command=lambda: self.mark_selected_students(listbox))
        mark_button.pack(pady=10)

    def mark_selected_students(self, listbox):
        # Get the selected students from the listbox
        selected_students = [listbox.get(idx) for idx in listbox.curselection()]

        # Mark attendance for selected students
        for student_name in selected_students:
            if not self.attendance_status[student_name]:
                self.attendance_status[student_name] = True
                attendance_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"Attendance Marked for {student_name} at {attendance_time}")

                # Call a function to mark attendance in the Excel file
                self.mark_attendance_in_excel(student_name, attendance_time)

        # Show a message box indicating successful attendance marking
        messagebox.showinfo("Attendance Marked", "Attendance marked for selected students.")

    def exit_app(self):
        # Release the webcam and close all windows
        self.video_capture.release()
        cv2.destroyAllWindows()
        # Destroy the GUI window
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = FaceRecognitionApp(root)
    root.mainloop()
